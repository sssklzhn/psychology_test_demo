# backend/parse_scales_correctly.py
import pandas as pd
import os

def parse_scales_correctly():
    """Правильный парсинг шкал из Excel файла"""
    
    excel_file = "psychological_test_questions.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Файл {excel_file} не найден!")
        return
    
    print("=== ПРАВИЛЬНЫЙ ПАРСИНГ ШКАЛ ===")
    
    # Сначала читаем ответы, чтобы знать какие вопросы имеют ответ "Да" или "Нет"
    df_answers = pd.read_excel(excel_file, sheet_name='Ответы', header=2)
    df_answers = df_answers.rename(columns={
        'Утверждения, относящиеся ко мне (моему характеру)': 'text',
        '№': 'number',
        'Да': 'yes',
        'Нет': 'no'
    })
    
    # Создаем словарь с информацией о вопросах
    questions_info = {}
    
    for idx, row in df_answers.iterrows():
        if pd.isna(row['number']):
            continue
        
        q_num = int(row['number'])
        has_yes = pd.notna(row.get('yes')) and row['yes'] == 1
        has_no = pd.notna(row.get('no')) and row['no'] == 1
        
        questions_info[q_num] = {
            'text': str(row['text']).strip(),
            'has_yes': has_yes,
            'has_no': has_no,
            'scales': []
        }
    
    print(f"Загружена информация о {len(questions_info)} вопросах")
    
    # Теперь читаем шкалы РУЧНО, анализируя структуру
    # На основе анализа Excel файла создаем маппинг вопросов к шкалам
    
    # В Excel файле в листе "Шкала" формулы вида:
    # =IF(Ответы!E4=1,1,0) - если в листе "Ответы" в столбце E (Нет) строка 4 стоит 1
    # =IF(Ответы!D5=1,1,0) - если в листе "Ответы" в столбце D (Да) строка 5 стоит 1
    
    # Создаем маппинг вручную на основе анализа файла
    # Это нужно сделать один раз, затем можно сохранить в JSON
    
    scales_mapping = {
        # Формат: номер_вопроса: [список_шкал]
        # Примеры из первых вопросов:
        1: ['Достоверность'],  # В Excel: =IF(Ответы!E4=1,1,0) - ответ "Нет" дает балл Достоверности
        2: ['Достоверность'],  # =IF(Ответы!D5=1,1,0) - ответ "Да" дает балл Достоверности
        3: ['Ранимость'],      # =IF(Ответы!D6=1,1,0) - ответ "Да" дает балл Ранимости
        4: ['Аутоагрессия', 'Истероидность', 'НПН'],  # 3 формулы
        5: ['Ранимость'],
        6: ['Достоверность'],
        7: ['Ранимость'],
        8: ['НПН'],
        9: ['Ранимость'],
        10: ['Ранимость', 'НПН'],
        11: ['Достоверность'],
        12: ['Ранимость'],
        13: ['Ранимость'],
        14: ['Ранимость'],
        15: ['НПН'],
        16: ['Ранимость'],
        17: ['Ранимость'],
        18: ['Достоверность'],
        19: ['Достоверность'],
        20: ['Аутоагрессия', 'Ранимость', 'НПН'],
        # ... и так далее для всех 160 вопросов
    }
    
    # Но так как делать вручную 160 вопросов долго, давайте создадим генератор
    print("\n=== АВТОМАТИЧЕСКИЙ АНАЛИЗ ШКАЛ ===")
    
    # Читаем сырые данные из листа "Шкала"
    df_scale_raw = pd.read_excel(excel_file, sheet_name='Шкала', header=None)
    
    # Найдем строку с заголовками шкал (строка 2, индексы: 0-верхний заголовок, 1-пустая, 2-заголовки шкал)
    print(f"Строка 2 (индекс 2): {df_scale_raw.iloc[2].tolist()}")
    
    # Вопросы начинаются с строки 3 (индекс 3 в pandas, но в Excel это строка 4)
    print("\nАнализ первых 20 вопросов:")
    
    scale_names = ['Isk', 'Con', 'Ast', 'Ist', 'Psi', 'NPN']
    scale_names_ru = {
        'Isk': 'Достоверность',
        'Con': 'Аутоагрессия',
        'Ast': 'Ранимость',
        'Ist': 'Истероидность',
        'Psi': 'Психопатическая',
        'NPN': 'НПН'
    }
    
    for i in range(3, 23):  # Строки 3-22 (первые 20 вопросов)
        try:
            question_num = df_scale_raw.iloc[i, 2]  # Столбец C
            if pd.isna(question_num):
                continue
            
            question_num = int(question_num)
            
            # Проверяем шкалы (столбцы D-I, индексы 3-8)
            question_scales = []
            
            for scale_idx, scale_col in enumerate(scale_names, start=3):
                cell_value = df_scale_raw.iloc[i, scale_idx]
                
                # В файле Excel формулы, но pandas их не читает
                # Вместо этого проверим - если в ячейке что-то есть (не NaN), значит вопрос относится к шкале
                if pd.notna(cell_value):
                    # Даже если это 0, в оригинале там формула
                    question_scales.append(scale_col)
            
            if question_scales:
                print(f"Вопрос {question_num}: {[scale_names_ru.get(s, s) for s in question_scales]}")
            
        except Exception as e:
            print(f"Ошибка в строке {i}: {e}")
    
    # Альтернативный подход: прочитать Excel с помощью openpyxl для получения формул
    print("\n=== ЧТЕНИЕ ФОРМУЛ С OPENPYXL ===")
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(excel_file, data_only=False)  # data_only=False чтобы получить формулы
        ws_scale = wb['Шкала']
        
        print(f"Лист 'Шкала' имеет размер: {ws_scale.max_row} строк, {ws_scale.max_column} столбцов")
        
        # Находим строки с вопросами (начиная с 4 строки Excel)
        for row in range(4, 24):  # Строки 4-23 Excel (первые 20 вопросов)
            question_num = ws_scale.cell(row=row, column=3).value  # Столбец C
            
            if question_num is None:
                continue
            
            print(f"\nВопрос {question_num}:")
            
            # Проверяем шкалы (столбцы D-I, колонки 4-9)
            for col in range(4, 10):
                cell = ws_scale.cell(row=row, column=col)
                if cell.value is not None:
                    # Получаем значение формулы
                    formula = cell.value
                    if isinstance(formula, str) and formula.startswith('='):
                        # Это формула!
                        scale_name = scale_names[col-4]  # 4->Isk, 5->Con, etc.
                        print(f"  {scale_names_ru[scale_name]}: {formula}")
                        
    except ImportError:
        print("openpyxl не установлен. Установите: pip install openpyxl")
    except Exception as e:
        print(f"Ошибка при чтении с openpyxl: {e}")

if __name__ == "__main__":
    parse_scales_correctly()